import json
import re
from pathlib import Path
import openpyxl
import lib_saberpro as sp

# Departamentos de la región Caribe para clasificación
DEPARTAMENTOS_CARIBE = {
    "ATLANTICO", "BOLIVAR", "CESAR", "CORDOBA", "LA GUAJIRA", "MAGDALENA", "SUCRE", 
    "SAN ANDRES", "SAN ANDRES Y PROVIDENCIA", "SAN ANDRES PROVIDENCIA Y SANTA CATALINA"
}

def load_excel_data(file_path: Path) -> list[dict[str, object]]:
    """
    Lee de forma perezosa el archivo Excel y retorna una lista de diccionarios
    solo con las filas que corresponden a las agregaciones necesarias.
    """
    print(f"Cargando {file_path.name}...")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    if 'SABER PRO' not in wb.sheetnames:
        print(f"Advertencia: No se encontró la hoja 'SABER PRO' en {file_path.name}")
        return []
        
    sheet = wb['SABER PRO']
    rows_iter = sheet.iter_rows(values_only=True)
    
    # Obtener cabecera y construir mapa de columnas
    try:
        header = [str(c).upper() for c in next(rows_iter)]
    except StopIteration:
        return []
        
    col_map = {col: idx for idx, col in enumerate(header)}
    
    # Validar que las columnas críticas existan
    criticas = ['AGREGACION', 'MEDIDA_AGREGACION', 'CANTIDADEVALUADOS', 'NOMBRE_INSTITUCION', 'NOMBRE_PRUEBA']
    for c in criticas:
        if c not in col_map:
            raise KeyError(f"Columna crítica '{c}' no encontrada en el archivo {file_path.name}")
            
    filtered_data = []
    
    # Agregaciones permitidas para procesar (limpias)
    agregaciones_permitidas = {"INSTITUCION", "PAIS", "DEPARTAMENTO", "PROGRAMA_ACADEMICO", "NBC"}
    
    for r in rows_iter:
        if len(r) <= max(col_map.values()):
            continue
            
        agreg = r[col_map['AGREGACION']]
        if agreg is None:
            continue
            
        agreg_clean = sp.clean_text(agreg)
        if agreg_clean not in agregaciones_permitidas:
            continue
            
        # Construir un objeto limpio para esta fila
        row_dict = {col: r[idx] for col, idx in col_map.items()}
        filtered_data.append(row_dict)
        
    wb.close()
    print(f"Filas filtradas leídas: {len(filtered_data)}")
    return filtered_data

def main() -> None:
    print("-------------------------------------------------------")
    params = sp.load_params()
    agregados_dir = sp.PROJECT_ROOT / params["agregados_dir"]
    anio_vigente = params["anio_vigente"]
    norm_mapping = sp.load_normalization()
    umbral_n_bajo = params.get("umbral_n_bajo_programa", 5)
    
    print(f"Iniciando procesamiento de bases agregadas (Fuente B)...")
    print(f"Directorio de entrada: {agregados_dir}")
    print(f"Año vigente: {anio_vigente}")
    print(f"Umbral para marcar programas con N bajo: {umbral_n_bajo}")
    
    if not agregados_dir.exists():
        raise FileNotFoundError(f"No existe el directorio de agregados: {agregados_dir}")
        
    excel_files = list(agregados_dir.glob("*.xlsx"))
    if not excel_files:
        raise FileNotFoundError(f"No se encontraron archivos Excel en {agregados_dir}")
        
    # Clasificación de programas de UNIMAGDALENA desde parametros.yml
    program_info = {}
    for p in params["programas_unimagdalena"]:
        clean_p_name = sp.clean_text(p["nombre"])
        program_info[clean_p_name] = {
            "facultad": p["facultad"],
            "nbc_id": p["nbc_id"],
            "nbc_nombre": p["nbc_nombre"]
        }
        
    sue_set = {sp.clean_text(name) for name in params["universidades_sue"]}
    caribe_set = {sp.clean_text(name) for name in params["universidades_caribe"]}
    competencias_gen = [sp.clean_text(c) for c in params["competencias_genericas"]]
    
    # Estructuras para acumular datos
    historico_institucional = []
    ranking_sue_2025 = []
    departamentos_2025 = []
    facultades_2025 = {}
    programas_2025 = {}
    top10_global_2025 = []
    top10_competencias_2025 = {c: [] for c in params["competencias_genericas"]}
    niveles_desempeno_2025 = []
    
    # Registro de cantidad de programas de UNIMAGDALENA por año para reporte
    conteo_programas_por_anio = {}
    alertas_n_bajo_2025 = []
    
    # Procesar cada archivo Excel
    for file in sorted(excel_files, key=lambda f: f.name):
        year_match = re.search(r"20\d{2}", file.name)
        if not year_match:
            print(f"Saltando archivo (no se detecta año en nombre): {file.name}")
            continue
        year = int(year_match.group())
        is_current = (year == anio_vigente)
        
        # Cargar los datos filtrados en memoria para el año
        rows = load_excel_data(file)
        if not rows:
            continue
            
        # 1. Procesamiento Institucional (UNIMAGDALENA) e Histórico
        unimag_global_score = None
        unimag_global_n = None
        national_global_score = None
        national_global_n = None
        
        # Competencias temporales para UNIMAGDALENA y Nacional
        unimag_comp_scores = {}
        national_comp_scores = {}
        
        # Extraer puntajes globales del país y de las IES
        for r in rows:
            agreg = sp.clean_text(r['AGREGACION'])
            medida = sp.clean_text(r['MEDIDA_AGREGACION'])
            inst = sp.normalize_ies_name(r['NOMBRE_INSTITUCION'], "agregados", norm_mapping)
            test_name = sp.clean_text(r['NOMBRE_PRUEBA'])
            
            # Global País
            if agreg == "PAIS" and medida == "PUNTAJE_GLOBAL":
                national_global_score = sp.safe_num(r['PROMEDIO_GLOBAL'])
                national_global_n = sp.safe_num(r['CANTIDADEVALUADOS'])
                
            # Global UNIMAGDALENA
            elif agreg == "INSTITUCION" and inst == "UNIVERSIDAD DEL MAGDALENA" and medida == "PUNTAJE_GLOBAL":
                unimag_global_score = sp.safe_num(r['PROMEDIO_GLOBAL'])
                unimag_global_n = sp.safe_num(r['CANTIDADEVALUADOS'])
                
            # Competencias País
            elif agreg == "PAIS" and medida == "PUNTAJE_PRUEBA" and test_name in competencias_gen:
                national_comp_scores[test_name] = {
                    "puntaje": sp.safe_num(r['PROMEDIO_PRUEBA']),
                    "n": sp.safe_num(r['CANTIDADEVALUADOS'])
                }
                
            # Competencias UNIMAGDALENA
            elif agreg == "INSTITUCION" and inst == "UNIVERSIDAD DEL MAGDALENA" and medida == "PUNTAJE_PRUEBA" and test_name in competencias_gen:
                unimag_comp_scores[test_name] = {
                    "puntaje": sp.safe_num(r['PROMEDIO_PRUEBA']),
                    "n": sp.safe_num(r['CANTIDADEVALUADOS'])
                }
                
        # Registrar histórico institucional
        if unimag_global_score is not None:
            historico_institucional.append({
                "anio": year,
                "puntaje_unimag": round(unimag_global_score, 2),
                "puntaje_nacional": round(national_global_score, 2) if national_global_score is not None else None,
                "n_unimag": int(unimag_global_n) if unimag_global_n is not None else None
            })
            
        # Conteo temporal de programas para reportar cantidad por año
        programas_anual_set = set()
        for r in rows:
            agreg = sp.clean_text(r['AGREGACION'])
            medida = sp.clean_text(r['MEDIDA_AGREGACION'])
            inst = sp.normalize_ies_name(r['NOMBRE_INSTITUCION'], "agregados", norm_mapping)
            prog_name = r['NOMBRE_PROGRAMA_ACAD']
            if agreg == "PROGRAMA_ACADEMICO" and inst == "UNIVERSIDAD DEL MAGDALENA" and prog_name is not None and medida == "PUNTAJE_GLOBAL":
                programas_anual_set.add(sp.clean_text(prog_name))
        conteo_programas_por_anio[str(year)] = len(programas_anual_set)
            
        # Si es el año vigente (2025), guardamos los detalles institucionales
        if is_current:
            institucional_global_2025 = {
                "puntaje_unimag": round(unimag_global_score, 2) if unimag_global_score is not None else None,
                "puntaje_nacional": round(national_global_score, 2) if national_global_score is not None else None,
                "n_unimag": int(unimag_global_n) if unimag_global_n is not None else 0,
                "n_nacional": int(national_global_n) if national_global_n is not None else 0
            }
            
            institucional_competencias_2025 = []
            for comp in params["competencias_genericas"]:
                comp_clean = sp.clean_text(comp)
                u_score = unimag_comp_scores.get(comp_clean, {}).get("puntaje")
                n_score = national_comp_scores.get(comp_clean, {}).get("puntaje")
                u_n = unimag_comp_scores.get(comp_clean, {}).get("n")
                
                institucional_competencias_2025.append({
                    "competencia": comp,
                    "puntaje_unimag": round(u_score, 2) if u_score is not None else None,
                    "puntaje_nacional": round(n_score, 2) if n_score is not None else None,
                    "n_unimag": int(u_n) if u_n is not None else 0
                })
                
            # 2. Ranking SUE 2025
            # Procesar las IES en el año vigente
            sue_dict = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                inst_raw = r['NOMBRE_INSTITUCION']
                
                if agreg == "INSTITUCION" and medida == "PUNTAJE_GLOBAL" and inst_raw is not None:
                    inst_norm = sp.normalize_ies_name(inst_raw, "agregados", norm_mapping)
                    clean_name = sp.clean_text(inst_norm)
                    
                    if clean_name in sue_set:
                        score = sp.safe_num(r['PROMEDIO_GLOBAL'])
                        n = sp.safe_num(r['CANTIDADEVALUADOS'])
                        if score is not None:
                            if inst_norm not in sue_dict or score > sue_dict[inst_norm]["puntaje"]:
                                sue_dict[inst_norm] = {
                                    "nombre": inst_norm,
                                    "puntaje": score,
                                    "n": n,
                                    "es_unimagdalena": (inst_norm == "UNIVERSIDAD DEL MAGDALENA"),
                                    "es_caribe": (clean_name in caribe_set)
                                }
            sorted_sue = sorted(sue_dict.values(), key=lambda x: x["puntaje"], reverse=True)
            for idx, item in enumerate(sorted_sue, 1):
                ranking_sue_2025.append({
                    "rank": idx,
                    "nombre": item["nombre"],
                    "puntaje": round(item["puntaje"], 2),
                    "n": int(item["n"]) if item["n"] is not None else 0,
                    "es_unimagdalena": item["es_unimagdalena"],
                    "es_caribe": item["es_caribe"]
                })
                
            # 3. Comparativo Departamental 2025
            dept_dict = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                dept = r['NOMBRE_DEPARTAMENTO']
                
                if agreg == "DEPARTAMENTO" and medida == "PUNTAJE_GLOBAL" and dept is not None:
                    dept_clean = sp.clean_text(dept)
                    score = sp.safe_num(r['PROMEDIO_GLOBAL'])
                    n = sp.safe_num(r['CANTIDADEVALUADOS'])
                    if score is not None:
                        dept_dict[dept_clean] = {
                            "departamento": dept_clean,
                            "puntaje": score,
                            "n": n,
                            "es_magdalena": (dept_clean == "MAGDALENA"),
                            "es_caribe": (dept_clean in DEPARTAMENTOS_CARIBE)
                        }
            sorted_depts = sorted(dept_dict.values(), key=lambda x: x["puntaje"], reverse=True)
            departamentos_2025 = [{
                "departamento": d["departamento"],
                "puntaje": round(d["puntaje"], 2),
                "n": int(d["n"]) if d["n"] is not None else 0,
                "es_magdalena": d["es_magdalena"],
                "es_caribe": d["es_caribe"]
            } for d in sorted_depts]
            
            # 4. Procesamiento de Programas 2025 contra su respectivo NBC
            # Primero recopilamos las medias nacionales globales de los NBCs
            nbc_global_nacionales = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                nbc_id = sp.safe_num(r['ID_NBC'])
                if agreg == "NBC" and medida == "PUNTAJE_GLOBAL" and nbc_id is not None:
                    score = sp.safe_num(r['PROMEDIO_GLOBAL'])
                    n_nbc = sp.safe_num(r['CANTIDADEVALUADOS'])
                    if score is not None:
                        nbc_global_nacionales[int(nbc_id)] = {
                            "puntaje": score,
                            "n": n_nbc
                        }
                        
            # Recopilamos las medias nacionales por prueba/competencia de los NBCs
            nbc_pruebas_nacionales = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                nbc_id = sp.safe_num(r['ID_NBC'])
                test_name = sp.clean_text(r['NOMBRE_PRUEBA'])
                
                if agreg == "NBC" and medida == "PUNTAJE_PRUEBA" and nbc_id is not None and test_name is not None:
                    score = sp.safe_num(r['PROMEDIO_PRUEBA'])
                    n_nbc = sp.safe_num(r['CANTIDADEVALUADOS'])
                    if score is not None:
                        nbc_pruebas_nacionales[(int(nbc_id), test_name)] = {
                            "puntaje": score,
                            "n": n_nbc
                        }
                        
            # Recopilamos niveles de desempeño por prueba/competencia de los NBCs nacionales
            nbc_niveles_nacionales = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                nbc_id = sp.safe_num(r['ID_NBC'])
                test_name = sp.clean_text(r['NOMBRE_PRUEBA'])
                
                if agreg == "NBC" and medida == "NIVEL_DESEMPENO_PRUEBA" and nbc_id is not None and test_name in competencias_gen:
                    n1 = sp.safe_num(r['NIVEL1'])
                    n2 = sp.safe_num(r['NIVEL2'])
                    n3 = sp.safe_num(r['NIVEL3'])
                    n4 = sp.safe_num(r['NIVEL4'])
                    n5 = sp.safe_num(r['NIVEL5'])
                    
                    levels = [n1, n2, n3, n4, n5]
                    levels = [float(l) if l is not None else 0.0 for l in levels]
                    nbc_niveles_nacionales[(int(nbc_id), test_name)] = levels
                    
            # Recopilar niveles de desempeño del país (nacionales) para genéricas (como fallback secundario)
            national_levels = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                test_name = sp.clean_text(r['NOMBRE_PRUEBA'])
                
                if agreg == "PAIS" and medida == "NIVEL_DESEMPENO_PRUEBA" and test_name in competencias_gen:
                    n1 = sp.safe_num(r['NIVEL1'])
                    n2 = sp.safe_num(r['NIVEL2'])
                    n3 = sp.safe_num(r['NIVEL3'])
                    n4 = sp.safe_num(r['NIVEL4'])
                    n5 = sp.safe_num(r['NIVEL5'])
                    
                    levels = [n1, n2, n3, n4, n5]
                    levels = [float(l) if l is not None else 0.0 for l in levels]
                    national_levels[test_name] = levels
                    
            # Recopilar niveles de desempeño de UNIMAGDALENA institucional para reporte general
            unimag_levels = {}
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                inst = sp.normalize_ies_name(r['NOMBRE_INSTITUCION'], "agregados", norm_mapping)
                test_name = sp.clean_text(r['NOMBRE_PRUEBA'])
                
                if agreg == "INSTITUCION" and inst == "UNIVERSIDAD DEL MAGDALENA" and medida == "NIVEL_DESEMPENO_PRUEBA" and test_name in competencias_gen:
                    n1 = sp.safe_num(r['NIVEL1'])
                    n2 = sp.safe_num(r['NIVEL2'])
                    n3 = sp.safe_num(r['NIVEL3'])
                    n4 = sp.safe_num(r['NIVEL4'])
                    n5 = sp.safe_num(r['NIVEL5'])
                    
                    levels = [n1, n2, n3, n4, n5]
                    levels = [float(l) if l is not None else 0.0 for l in levels]
                    unimag_levels[test_name] = levels
                    
            for comp in params["competencias_genericas"]:
                comp_clean = sp.clean_text(comp)
                niveles_desempeno_2025.append({
                    "competencia": comp,
                    "distribucion_unimag": unimag_levels.get(comp_clean, [0.0, 0.0, 0.0, 0.0, 0.0]),
                    "distribucion_nacional": national_levels.get(comp_clean, [0.0, 0.0, 0.0, 0.0, 0.0])
                })
                
            # Procesar programas de UNIMAGDALENA para el año vigente (2025)
            # Primero leemos puntaje global de cada programa y lo cruzamos con el NBC global nacional de referencia
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                inst = sp.normalize_ies_name(r['NOMBRE_INSTITUCION'], "agregados", norm_mapping)
                prog_name = r['NOMBRE_PROGRAMA_ACAD']
                
                if agreg == "PROGRAMA_ACADEMICO" and inst == "UNIVERSIDAD DEL MAGDALENA" and prog_name is not None:
                    prog_clean = sp.clean_text(prog_name)
                    
                    if prog_clean in program_info:
                        info = program_info[prog_clean]
                        fac = info["facultad"]
                        nbc_id = info["nbc_id"]
                        
                        if prog_clean not in programas_2025:
                            programas_2025[prog_clean] = {
                                "programa": prog_name,
                                "facultad": fac,
                                "nbc_id": nbc_id,
                                "nbc_nombre": info["nbc_nombre"],
                                "global": None,
                                "global_nbc_nacional": None,
                                "n_nbc_nacional": None,
                                "n": 0,
                                "competencias": {},
                                "especificas": {},
                                "niveles": {},
                                "historico": []
                            }
                            
                        if medida == "PUNTAJE_GLOBAL":
                            score = sp.safe_num(r['PROMEDIO_GLOBAL'])
                            n = sp.safe_num(r['CANTIDADEVALUADOS'])
                            
                            # Obtener referencia global del NBC
                            nbc_ref_global = nbc_global_nacionales.get(int(nbc_id), {})
                            nbc_score_global = nbc_ref_global.get("puntaje")
                            nbc_n_global = nbc_ref_global.get("n")
                            
                            programas_2025[prog_clean]["global"] = round(score, 2) if score is not None else None
                            programas_2025[prog_clean]["n"] = int(n) if n is not None else 0
                            programas_2025[prog_clean]["global_nbc_nacional"] = round(nbc_score_global, 2) if nbc_score_global is not None else None
                            programas_2025[prog_clean]["n_nbc_nacional"] = int(nbc_n_global) if nbc_n_global is not None else None
                            
                            # Guardar alertas de N bajo sin abortar
                            if n is not None and n < umbral_n_bajo:
                                alertas_n_bajo_2025.append(f"{prog_name} (n={n})")
                                
            # Recopilar las pruebas específicas y genéricas por programa de UNIMAGDALENA y sus niveles
            for r in rows:
                agreg = sp.clean_text(r['AGREGACION'])
                medida = sp.clean_text(r['MEDIDA_AGREGACION'])
                inst = sp.normalize_ies_name(r['NOMBRE_INSTITUCION'], "agregados", norm_mapping)
                prog_name = r['NOMBRE_PROGRAMA_ACAD']
                test_name = r['NOMBRE_PRUEBA']
                
                if agreg == "PROGRAMA_ACADEMICO" and inst == "UNIVERSIDAD DEL MAGDALENA" and prog_name is not None and test_name is not None:
                    prog_clean = sp.clean_text(prog_name)
                    test_clean = sp.clean_text(test_name)
                    
                    if prog_clean in programas_2025:
                        prog_obj = programas_2025[prog_clean]
                        nbc_id = prog_obj["nbc_id"]
                        
                        # Buscar puntaje de la prueba
                        if medida == "PUNTAJE_PRUEBA":
                            score = sp.safe_num(r['PROMEDIO_PRUEBA'])
                            if score is not None:
                                # Buscar promedio nacional de referencia del NBC, no de PAIS
                                nbc_ref = nbc_pruebas_nacionales.get((int(nbc_id), test_clean), {})
                                nbc_score = nbc_ref.get("puntaje")
                                nbc_n = nbc_ref.get("n")
                                
                                test_obj = {
                                    "prueba": test_name,
                                    "puntaje_programa": round(score, 2),
                                    "puntaje_nbc_nacional": round(nbc_score, 2) if nbc_score is not None else None,
                                    "n_nbc_nacional": int(nbc_n) if nbc_n is not None else None
                                }
                                
                                if test_clean in competencias_gen:
                                    prog_obj["competencias"][test_clean] = test_obj
                                else:
                                    prog_obj["especificas"][test_clean] = test_obj
                                    
                        # Buscar niveles de desempeño (comparados con su NBC, no con PAIS)
                        elif medida == "NIVEL_DESEMPENO_PRUEBA" and test_clean in competencias_gen:
                            n1 = sp.safe_num(r['NIVEL1'])
                            n2 = sp.safe_num(r['NIVEL2'])
                            n3 = sp.safe_num(r['NIVEL3'])
                            n4 = sp.safe_num(r['NIVEL4'])
                            n5 = sp.safe_num(r['NIVEL5'])
                            
                            levels = [n1, n2, n3, n4, n5]
                            levels = [float(l) if l is not None else 0.0 for l in levels]
                            
                            # Buscar niveles del NBC nacional
                            nbc_levels = nbc_niveles_nacionales.get((int(nbc_id), test_clean))
                            if nbc_levels is None:
                                # Fallback a país en caso de no encontrarse para el NBC
                                nbc_levels = national_levels.get(test_clean, [0.0, 0.0, 0.0, 0.0, 0.0])
                                
                            prog_obj["niveles"][test_clean] = {
                                "distribucion_programa": levels,
                                "distribucion_nbc_nacional": nbc_levels
                            }

    # 5. Rellenar histórico de programas leyendo todos los años de forma dinámica
    print("Construyendo históricos por programa...")
    for file in sorted(excel_files, key=lambda f: f.name):
        year_match = re.search(r"20\d{2}", file.name)
        if not year_match:
            continue
        year = int(year_match.group())
        
        # Leemos solo filas de programa para optimizar la carga histórica
        wb = openpyxl.load_workbook(file, read_only=True)
        sheet = wb['SABER PRO']
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(c).upper() for c in next(rows_iter)]
        col_map = {col: idx for idx, col in enumerate(header)}
        
        for r in rows_iter:
            if len(r) <= max(col_map.values()):
                continue
            agreg = sp.clean_text(r[col_map['AGREGACION']])
            medida = sp.clean_text(r[col_map['MEDIDA_AGREGACION']])
            inst = sp.normalize_ies_name(r[col_map['NOMBRE_INSTITUCION']], "agregados", norm_mapping)
            prog_name = r[col_map['NOMBRE_PROGRAMA_ACAD']]
            
            if agreg == "PROGRAMA_ACADEMICO" and inst == "UNIVERSIDAD DEL MAGDALENA" and medida == "PUNTAJE_GLOBAL" and prog_name is not None:
                prog_clean = sp.clean_text(prog_name)
                if prog_clean in programas_2025:
                    score = sp.safe_num(r[col_map['PROMEDIO_GLOBAL']])
                    n = sp.safe_num(r[col_map['CANTIDADEVALUADOS']])
                    if score is not None:
                        programas_2025[prog_clean]["historico"].append({
                            "anio": year,
                            "puntaje": round(score, 2),
                            "n": int(n) if n is not None else 0
                        })
        wb.close()

    # Ordenar los históricos cronológicamente
    for prog_clean in programas_2025:
        programas_2025[prog_clean]["historico"] = sorted(
            programas_2025[prog_clean]["historico"], 
            key=lambda x: x["anio"]
        )

    # 6. Agregación de Facultades (año vigente - 2025)
    # Calculamos el promedio ponderado de los programas de UNIMAGDALENA
    print("Agregando resultados por Facultad para 2025...")
    for prog_clean, p in programas_2025.items():
        fac = p["facultad"]
        n_prog = p["n"]
        score_prog = p["global"]
        
        if fac not in facultades_2025:
            facultades_2025[fac] = {
                "facultad": fac,
                "puntaje_global": 0.0,
                "n": 0,
                "temp_score_sum": 0.0,
                "competencias_sum": {sp.clean_text(c): 0.0 for c in params["competencias_genericas"]},
                "competencias_n": {sp.clean_text(c): 0 for c in params["competencias_genericas"]}
            }
            
        fac_obj = facultades_2025[fac]
        if score_prog is not None and n_prog > 0:
            fac_obj["n"] += n_prog
            fac_obj["temp_score_sum"] += score_prog * n_prog
            
        for test_clean, comp_obj in p["competencias"].items():
            comp_score = comp_obj["puntaje_programa"]
            if comp_score is not None and n_prog > 0:
                fac_obj["competencias_sum"][test_clean] += comp_score * n_prog
                fac_obj["competencias_n"][test_clean] += n_prog
                
    # Finalizar cálculo de promedios de facultades
    facultades_salida_2025 = []
    for fac, f in facultades_2025.items():
        score_global = f["temp_score_sum"] / f["n"] if f["n"] > 0 else None
        
        competencias_salida = []
        for comp in params["competencias_genericas"]:
            comp_clean = sp.clean_text(comp)
            sum_score = f["competencias_sum"].get(comp_clean, 0.0)
            n_sum = f["competencias_n"].get(comp_clean, 0)
            comp_avg = sum_score / n_sum if n_sum > 0 else None
            
            competencias_salida.append({
                "competencia": comp,
                "puntaje": round(comp_avg, 2) if comp_avg is not None else None
            })
            
        facultades_salida_2025.append({
            "facultad": fac,
            "puntaje_global": round(score_global, 2) if score_global is not None else None,
            "n": f["n"],
            "competencias": competencias_salida
        })
        
    # 7. Construcción de Top 10 para 2025
    list_progs = [p for p in programas_2025.values() if p["global"] is not None]
    top10_global_2025 = sorted(list_progs, key=lambda x: x["global"], reverse=True)[:10]
    top10_global_salida = [{
        "programa": p["programa"],
        "puntaje": p["global"],
        "n": p["n"]
    } for p in top10_global_2025]
    
    top10_competencias_salida = {}
    for comp in params["competencias_genericas"]:
        comp_clean = sp.clean_text(comp)
        prog_comp_list = []
        for p in programas_2025.values():
            if comp_clean in p["competencias"]:
                score = p["competencias"][comp_clean]["puntaje_programa"]
                if score is not None:
                    prog_comp_list.append({
                        "programa": p["programa"],
                        "puntaje": score
                    })
        sorted_comp_list = sorted(prog_comp_list, key=lambda x: x["puntaje"], reverse=True)[:10]
        top10_competencias_salida[comp] = sorted_comp_list
        
    # Convertir a lista limpia para salida JSON
    programas_salida_2025 = []
    for p in programas_2025.values():
        programas_salida_2025.append({
            "programa": p["programa"],
            "facultad": p["facultad"],
            "n_2025": p["n"],
            "global_2025": p["global"],
            "global_nbc_nacional_2025": p.get("global_nbc_nacional"),
            "n_nbc_nacional_2025": p.get("n_nbc_nacional"),
            "n_bajo": p["n"] < umbral_n_bajo,
            "competencias_2025": list(p["competencias"].values()),
            "especificas_2025": list(p["especificas"].values()),
            "niveles_2025": [
                {
                    "competencia": next(c for c in params["competencias_genericas"] if sp.clean_text(c) == comp_clean),
                    "distribucion_programa": info["distribucion_programa"],
                    "distribucion_nbc_nacional": info["distribucion_nbc_nacional"]
                }
                for comp_clean, info in p["niveles"].items()
            ],
            "historico": p["historico"]
        })
        
    # Guardar salida agregada
    output_data = {
        "institucional": {
            "global": institucional_global_2025,
            "competencias": institucional_competencias_2025,
            "historico": sorted(historico_institucional, key=lambda x: x["anio"])
        },
        "sue_ranking": ranking_sue_2025,
        "departamento": departamentos_2025,
        "facultades": sorted(facultades_salida_2025, key=lambda x: x["facultad"]),
        "programas": sorted(programas_salida_2025, key=lambda x: x["programa"]),
        "top10": {
            "global": top10_global_salida,
            "por_competencia": top10_competencias_salida
        },
        "niveles_desempeno": niveles_desempeno_2025
    }
    
    processed_dir = sp.PROJECT_ROOT / params["processed_dir"]
    processed_dir.mkdir(parents=True, exist_ok=True)
    output_path = processed_dir / "agregados_procesados.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
        
    print(f"Procesamiento finalizado. Archivo temporal guardado en {output_path}")
    print(f"Años procesados para histórico: {len(excel_files)}")
    print(f"Conteo dinámico de programas por año: {conteo_programas_por_anio}")
    print(f"Total programas de UNIMAGDALENA en 2025 (deduplicados): {len(programas_salida_2025)}")
    print(f"Programas con N bajo en 2025 (n < {umbral_n_bajo}): {len(alertas_n_bajo_2025)}")
    for alert in sorted(alertas_n_bajo_2025):
        print(f"  - Alerta N bajo: {alert}")
    print("-------------------------------------------------------")

if __name__ == "__main__":
    main()
